#!/usr/bin/env python
# -*- coding: utf-8 -*-
# Created by Jiawei Liu on 2018/06/18
import os
import re

import json
from openpyxl import load_workbook
import nltk
import numpy as np
import pandas as pd
import tensorflow as tf
from tensorflow.python.estimator.estimator import Estimator
import yaml
from spacy.lang.en.stop_words import STOP_WORDS
import spacy

from bert import tokenization, modeling
from bert.extract_features import model_fn_builder, convert_lst_to_features, PoolingStrategy
from bert_serving.client import BertClient

# 加载打分系统配置， 主要是系列模型文件的路径
with open("config/sys_conf.yaml", encoding="utf-8") as conf_reader:
    sys_conf = yaml.load(conf_reader.read())

# do_train: Whether to run training.
# do_eval: Whether to run eval on the dev set.
# do_predict: Whether to run the model in inference mode on the test set.
# train_batch_size: Total batch size for training.
# eval_batch_size: Total batch size for eval.
# predict_batch_size: Total batch size for predict.
# learning_rate: The initial learning rate for Adam.
# num_train_epochs: Total number of training epochs to perform.
# warmup_proportion: Proportion of training to perform linear learning rate warmup for.
# save_checkpoints_step: How often to save the model checkpoint.
# iterations_per_loop: How many steps to make in each estimator call.
# prompt_id: the id of the prompt
# train_set_prob: the Proportion examples from dataset chosen to be the train set
with open("config/train_conf.json", "r") as cr:
    train_conf = json.load(cr)

with open("config/doc_conf.json", "r") as cr:
    doc_conf = json.load(cr)

spacynlp = spacy.load("en_core_web_sm")

# 目前所有数据集的统计数据， key代表了prompt的id, 目前1-8是asap数据集的id, 9代表雅思，
# 之后积累数据后，会把雅思的每个prompt都用一个id表示，id当初后面的三元组(最低分，最高分，该prompt下的范文总数)
dataset_score_range = {
    1: (2, 12, 1783),
    2: (1, 6, 1800),
    3: (0, 3, 1726),
    4: (0, 3, 1772),
    5: (0, 4, 1805),
    6: (0, 4, 1800),
    7: (0, 30, 1569),
    8: (0, 60, 723),
    9: (0, 9, 429)
}


class ScoreResource:
    """ 整个score系统需要

    Attributes:
        advanced_vocabulary_name:

    """

    def __init__(self):
        self.advanced_vocabulary_name = ["5.5", "5.5-6.5", "6.5-7.5"]
        self.advanced_words = self.__load_ad_words(sys_conf["advanced_word_path"])

    def __load_ad_words(self, ad_word_path):
        """ 加载高级词汇，5.5-7分的词汇等，之后会进行一定修改

        Args:
            ad_word_path: 存放高级词汇文件的路径

        Returns:
            ad_word_dict: 字典对象，key: 单词，value:单词所属的分数段

        """
        assert os.path.exists(ad_word_path), "advance words file path is not exists, please check score_conf.yaml"
        ad_word_dict = dict()
        for dirpath, dirnames, filenames in os.walk(ad_word_path):
            for filename in filenames:
                if filename[-4:] == 'xlsx' and filename[:-5] in self.advanced_vocabulary_name:
                    try:
                        filepath = os.path.join(dirpath, filename)
                        word_file = load_workbook(filepath)
                        sheet = word_file[word_file.sheetnames[0]]
                        rows = sheet.rows
                        for row in rows:
                            word = [col.value for col in row][0]
                            if not ad_word_dict.__contains__(word):
                                ad_word_dict[word] = filename[:-5]
                            else:
                                if float(filename[:-5].split('-')[-1]) > float(ad_word_dict[word].split('-')[-1]):
                                    ad_word_dict[word] = filename[:-5]
                    except:
                        raise ValueError("loading words .xlsx file error.")
        return ad_word_dict


sr = ScoreResource()


class Document:
    """ 输入给score系统的打分的对象，会根据gec传入的结果计算一篇文章的handmarked的特征

    Attributes:
        __title: 作文题目， 用spacy包装的对象
        __essay: 作文正文， list对象，其中每个元素都是对应sentence的spacy封装的对象
        __gec_output: gec的结果

    """

    def __init__(self, gec_output):
        """

        Args:
            gec_output: 由json.loads起来的对象，fupugec-server项目中传过来的对象。

        """
        self.__title = spacynlp(gec_output["title"].lower())
        self.__essay = [spacynlp(gec_output["sentence_" + str(index)]["orig_sent"].lower()) for index in
                        range(int(gec_output["sent_nums"]))]
        self.__gec_output = gec_output
        self.features = self.__features()
        self.__ad_vocab = self.__advanced_vocab()
        self.doc_result = self.__doc_result()

    def __doc_info(self):
        """ 文章总体的一些指标，

        Returns: list对象， 包括[总词数，总字符数，平均词长，词汇数，词长方差，介词数，句数，平均句长，句长方差]

        """
        process_sent = lambda sent: [(token.text, token.lemma_) for token in sent if
                                     not (token.is_punct or token.is_space)]
        doc_token_and_lemma = []
        doc_sent_word_leng = []
        # word level
        self.__doc_num_short_sents = 0
        self.__doc_num_long_sents = 0
        for sent in self.__essay:
            temp_process_sent = process_sent(sent)
            if len(temp_process_sent) <= doc_conf["num_short_sentence_word"]:
                self.__doc_num_short_sents += 1
            if len(temp_process_sent) >= doc_conf["num_long_sentence_word"]:
                self.__doc_num_long_sents += 1
            doc_token_and_lemma.extend(temp_process_sent)
            doc_sent_word_leng.append(len(temp_process_sent))

        self.__doc_num_words = len(doc_token_and_lemma)  # 文章单词的数目
        doc_words_char_leng = [len(item[0]) for item in doc_token_and_lemma]  # 文章单词的平均长度
        doc_num_characters = sum(doc_words_char_leng)  # 文章总得character数目
        doc_average_word_leng = np.mean(doc_words_char_leng)
        doc_var_word_leng = np.var(doc_words_char_leng)

        self.__doc_vocab = set([item[1] for item in doc_token_and_lemma]) - STOP_WORDS
        self.__doc_num_vocab = len(self.__doc_vocab)

        prepositions = []
        preposition_detect = lambda sent: [token for token in sent if token.dep_ == "prep"]
        for sent in self.__essay:
            temp_prepostion = preposition_detect(sent)
            prepositions.extend(temp_prepostion)
        doc_num_prepositions = len(prepositions)

        # sent level
        self.__doc_num_sents = len(doc_sent_word_leng)  # 句子数目
        doc_average_sent_leng = np.mean(doc_sent_word_leng)  # 句子平均单词数目
        doc_var_sent_leng = np.var(doc_sent_word_leng)  # 句子单词数的方差
        clause_sent_num = 0
        for sentence_index in range(int(self.__gec_output['sent_nums'])):
            sentence = self.__gec_output["sentence_" + str(sentence_index)]
            # -1 表示不是从句， 7表示there be句型
            if not sentence["sent_type"] in [-1, 7]:
                clause_sent_num += 1

        # doc level
        self.__doc_num_paras = self.__gec_output["para_nums"]

        return [self.__doc_num_words, doc_num_characters, doc_average_word_leng, self.__doc_num_vocab,
                doc_var_word_leng,
                doc_num_prepositions,
                self.__doc_num_sents, doc_average_sent_leng, doc_var_sent_leng]

    def __error_info(self):
        """ 文章的错误信息，主要是错词率和错句率，这里没有使用具体数量，因为长文章倾向于错词数肯定要大于短文章。

        Returns: list对象，[ 错词率，错句率]

        """
        self.__err_word = list()
        doc_num_err = 0
        self.__doc_num_err_sentence = 0
        err_sentence_sign = False

        for sentence_index in range(int(self.__gec_output["sent_nums"])):
            sentence = self.__gec_output["sentence_" + str(sentence_index)]
            if sentence["err_num"] == 0:
                continue
            else:
                for edit_index in range(sentence["err_num"]):
                    edit = sentence["edit_" + str(edit_index)]
                    if edit["err_type"][2:] == "SPELL":
                        self.__err_word.append({
                            "err_word": sentence["orig_sent"].strip().split()[edit["start_err"]],
                            "corr_word": edit["corr_str"]
                        })
                    if not edit["err_type"][2:] in ["ORTH", "PUNCT"]:
                        doc_num_err += 1
                        err_sentence_sign = True
                if err_sentence_sign:
                    self.__doc_num_err_sentence += 1
                    err_sentence_sign = False
        return [doc_num_err / self.__doc_num_words, self.__doc_num_err_sentence / self.__doc_num_sents]

    def __advanced_vocab(self):
        """ 统计文章高级词汇

        Returns: 文章高级词汇字典，key->分数段，value->单词列表

        """
        ad_vocab = dict()
        for key in sr.advanced_vocabulary_name:
            ad_vocab[key] = list()
        for word in self.__doc_vocab:
            if sr.advanced_words.__contains__(word):
                ad_vocab[sr.advanced_words[word]].append(word)
        return ad_vocab

    def __features(self):
        """ 文章所有的

        Returns: 一篇文章的handmarked的特征集合

        """
        feature_list = self.__doc_info()
        feature_list.extend(self.__error_info())
        return feature_list

    def __doc_result(self):
        """ 生成需要输出报告的文章属性字典

        Returns: dict, document需要输出的属性

        """
        result = dict()
        result["num_word"] = self.__doc_num_words
        result["num_sentence"] = self.__doc_num_sents
        result['num_short_sentence'] = self.__doc_num_short_sents  # 短句数
        result['num_long_sentence'] = self.__doc_num_long_sents  # 长句数
        result['num_paragraph'] = self.__doc_num_paras  # 段落数

        result["err_word"] = self.__err_word
        result["num_err_word"] = len(self.__err_word)
        result['num_err_sentence'] = self.__doc_num_err_sentence  # 出现错误的句数

        result[
            'ratio_short_sentence'] = self.__doc_num_short_sents * 1.0 / self.__doc_num_sents if self.__doc_num_sents != 0 else 0  # 短句占比
        result[
            'ratio_long_sentence'] = self.__doc_num_long_sents * 1.0 / self.__doc_num_sents if self.__doc_num_sents != 0 else 0  # 长句占比
        result['err_rate_word'] = len(
            self.__err_word) * 1.0 / self.__doc_num_vocab if self.__doc_num_vocab != 0 else 0  # 错词占比
        result[
            'err_rate_sentence'] = self.__doc_num_err_sentence * 1.0 / self.__doc_num_sents if self.__doc_num_sents != 0 else 0  # 错句占比

        result['word_5.5'] = self.__ad_vocab['5.5']  # 文中属于5.5分的词汇列表
        result['word_5.5-6.5'] = self.__ad_vocab['5.5-6.5']  # 文中属于5.5-6.5分的词汇列表
        result['word_6.5-7.5'] = self.__ad_vocab['6.5-7.5']  # 文中属于5.5分的词汇列表

        return result


class BertWorker:
    def __init__(self):
        # the pooling layer index of bert-original model
        self.pooling_layer = [-2]
        # the pooling_strategy of bert-original model
        self.pooling_strategy = PoolingStrategy.REDUCE_MEAN
        # "The maximum total input sequence length after WordPiece tokenization. "
        # "Sequences longer than this will be truncated, and sequences shorter "
        # "than this will be padded."
        self.max_seq_len = 128

        self.bert_model_dir = sys_conf["bert_dir"]
        self.config_fp = os.path.join(self.bert_model_dir, "bert_config.json")
        self.ckpt_fp = os.path.join(self.bert_model_dir, "bert_model.ckpt")
        self.vocab_fp = os.path.join(self.bert_model_dir, "vocab.txt")
        self.tokenizer = tokenization.FullTokenizer(vocab_file=self.vocab_fp)
        self.model_fn = model_fn_builder(
            bert_config=modeling.BertConfig.from_json_file(self.config_fp),
            init_checkpoint=self.ckpt_fp,
            pooling_strategy=self.pooling_strategy,
            pooling_layer=self.pooling_layer
        )
        self.estimator = Estimator(self.model_fn)

    def input_fn_builder_file_path(self, file_path):
        def gen_asap_article():
            dataset = pd.read_csv(file_path)
            articles = dataset["essay"]
            articles_set = dataset["essay_set"]
            domain1_score = dataset["domain1_score"]
            articles_id = dataset["essay_id"]
            for i in range(len(articles)):
                doc = articles[i]
                sentences = sentence_tokenize(doc)
                tmp_f = list(convert_lst_to_features(sentences, self.max_seq_len, self.tokenizer))
                yield {
                    "input_ids": [f.input_ids for f in tmp_f],
                    "input_mask": [f.input_mask for f in tmp_f],
                    "input_type_ids": [f.input_type_ids for f in tmp_f],
                    "article_set": articles_set[i],
                    "domain1_score": float(domain1_score[i]),
                    "article_id": articles_id[i]
                }

        def input_fn():
            return (tf.data.Dataset.from_generator(
                gen_asap_article,
                output_types={
                    "input_ids": tf.int32,
                    "input_mask": tf.int32,
                    "input_type_ids": tf.int32,
                    "article_set": tf.int32,
                    "domain1_score": tf.float32,
                    "article_id": tf.int32
                },
                output_shapes={
                    "input_ids": (None, self.max_seq_len),
                    "input_mask": (None, self.max_seq_len),
                    "input_type_ids": (None, self.max_seq_len),
                    "article_set": [],
                    "domain1_score": [],
                    "article_id": []
                }
            ))

        return input_fn

    def inference_from_path_with_permfile(self, file_path):
        input_fn = self.input_fn_builder_file_path(file_path)
        for r in self.estimator.predict(input_fn, yield_single_examples=False):
            temp_sample = {"doc_encodes": r["encodes"],
                           "article_set": r["article_set"],
                           "domain1_score": r["domain1_score"],
                           "article_id": r["article_id"]}
            yield temp_sample

    def input_fn_builder_eilts_path(self, essay_path, score_path):
        def gen_eilts_article():
            score = dict()
            with open(score_path, "r", encoding="utf-8") as sr:
                for line in sr:
                    score[line.split()[0]] = float(line.split()[1])

            for dirpath, dirnames, filenames in os.walk(essay_path):
                if filenames:
                    for filename in filenames:
                        filepath = os.path.join(dirpath, filename)
                        with open(filepath, "r") as dr:
                            lines = []
                            for line in dr:
                                if line.strip():
                                    lines.append(line.strip())
                            title_and_doc = " ".join(lines)
                            title = title_and_doc.split("\t", 1)[0].strip()
                            doc = title_and_doc.split("\t", 1)[1].strip()
                            sentences = sentence_tokenize(doc)
                            tmp_f = list(convert_lst_to_features(sentences, self.max_seq_len, self.tokenizer))
                            yield {
                                "input_ids": [f.input_ids for f in tmp_f],
                                "input_mask": [f.input_mask for f in tmp_f],
                                "input_type_ids": [f.input_type_ids for f in tmp_f],
                                "article_set": 9,
                                "domain1_score": float(score[filename]),
                                "article_id": int(filename)
                            }

        def input_fn():
            return (tf.data.Dataset.from_generator(
                gen_eilts_article,
                output_types={
                    "input_ids": tf.int32,
                    "input_mask": tf.int32,
                    "input_type_ids": tf.int32,
                    "article_set": tf.int32,
                    "domain1_score": tf.float32,
                    "article_id": tf.int32
                },
                output_shapes={
                    "input_ids": (None, self.max_seq_len),
                    "input_mask": (None, self.max_seq_len),
                    "input_type_ids": (None, self.max_seq_len),
                    "article_set": [],
                    "domain1_score": [],
                    "article_id": []
                }
            ))

        return input_fn

    def inference_from_eitls_path(self, essay_path, score_path):
        input_fn = self.input_fn_builder_eilts_path(essay_path, score_path)
        for r in self.estimator.predict(input_fn, yield_single_examples=False):
            temp_sample = {"doc_encodes": r["encodes"],
                           "article_set": r["article_set"],
                           "domain1_score": r["domain1_score"],
                           "article_id": r["article_id"]}
            yield temp_sample

    def input_fn_builder_client(self):

        pass

    def inference_from_client(self):
        pass


def read_dataset_into_tfrecord(dataset_path, bw: BertWorker):
    dataset_positive_path = os.path.join(dataset_path, "training_set_rel3.csv")
    tf_record_path = os.path.join(dataset_path, "asap_dataset.tfrecord")

    # TODO(Jiawei):提取公共代码
    with tf.python_io.TFRecordWriter(tf_record_path) as tfrecord_writer:
        for i, item in enumerate(bw.inference_from_path_with_permfile(dataset_positive_path)):
            if i % 100 == 0:
                tf.logging.info("process {} docs".format(i))
            features = {}
            features["doc_encodes"] = tf.train.Feature(
                float_list=tf.train.FloatList(value=item["doc_encodes"].reshape(-1)))
            features["doc_encodes_shape"] = tf.train.Feature(
                int64_list=tf.train.Int64List(value=item["doc_encodes"].shape))
            features["article_set"] = tf.train.Feature(int64_list=tf.train.Int64List(value=[item["article_set"]]))
            features["article_id"] = tf.train.Feature(int64_list=tf.train.Int64List(value=[item["article_id"]]))
            features["domain1_score"] = tf.train.Feature(float_list=tf.train.FloatList(value=[item["domain1_score"]]))
            features["doc_sent_num"] = tf.train.Feature(
                int64_list=tf.train.Int64List(value=[item["doc_encodes"].shape[0]]))
            tf_features = tf.train.Features(feature=features)
            tf_example = tf.train.Example(features=tf_features)
            tfrecord_writer.write(tf_example.SerializeToString())


def read_ielts_into_tfrecord(dataset_path, bw: BertWorker):
    essay_path = os.path.join(dataset_path, "essay")
    score_path = os.path.join(dataset_path, "score")
    tf_record_path = os.path.join(dataset_path, "ielts.tfrecord")

    with tf.python_io.TFRecordWriter(tf_record_path) as tfrecord_writer:
        for i, item in enumerate(bw.inference_from_eitls_path(essay_path, score_path)):
            if i % 100 == 0:
                tf.logging.info("process {} docs".format(i))
            features = {}
            features["doc_encodes"] = tf.train.Feature(
                float_list=tf.train.FloatList(value=item["doc_encodes"].reshape(-1)))
            features["doc_encodes_shape"] = tf.train.Feature(
                int64_list=tf.train.Int64List(value=item["doc_encodes"].shape))
            features["article_set"] = tf.train.Feature(int64_list=tf.train.Int64List(value=[item["article_set"]]))
            features["article_id"] = tf.train.Feature(int64_list=tf.train.Int64List(value=[item["article_id"]]))
            features["domain1_score"] = tf.train.Feature(float_list=tf.train.FloatList(value=[item["domain1_score"]]))
            features["doc_sent_num"] = tf.train.Feature(
                int64_list=tf.train.Int64List(value=[item["doc_encodes"].shape[0]]))
            tf_features = tf.train.Features(feature=features)
            tf_example = tf.train.Example(features=tf_features)
            tfrecord_writer.write(tf_example.SerializeToString())


def create_initializer(initializer_range=0.02):
    """ 创建tensorflow初始化器

    Args:
        initializer_range: 初始化的范围设置

    Returns:

    """
    return tf.truncated_normal_initializer(stddev=initializer_range)


def create_rnn_cell(hidden_size, dropout_prob, layers_num, isbidirectional, is_training):
    """ 创建rnn cell, 包括多层和单双向的控制。

    Args:
        hidden_size: rnn cell 隐层的宽度
        dropout_prob: dropout的比例
        layers_num: rnn 网络的层数
        isbidirectional: 是否使用双向的cell
        is_training: 是否为训练时间段，训练时间段，cell需要使用dropout包裹

    Returns: cell实例

    """

    def single_rnn_cell():
        single_cell = tf.nn.rnn_cell.LSTMCell(hidden_size)
        if is_training:
            single_cell = tf.nn.rnn_cell.DropoutWrapper(single_cell, output_keep_prob=1 - dropout_prob)
        return single_cell

    if isbidirectional:
        fw_cell = tf.nn.rnn_cell.MultiRNNCell([single_rnn_cell() for _ in range(layers_num)])
        bw_cell = tf.nn.rnn_cell.MultiRNNCell([single_rnn_cell() for _ in range(layers_num)])
        cell = [fw_cell, bw_cell]
    else:
        fw_cell = tf.nn.rnn_cell.MultiRNNCell([single_rnn_cell() for _ in range(layers_num)])
        cell = fw_cell
    return cell


def input_fn_from_client(bw: BertWorker):
    def gen():
        for item in bw.inference_from_client():
            temp_sample = {
                "doc_encodes": item["doc_encodes"],
                "article_set": item["article_set"],
                "domain1_score": item["domain1_score"],
                "article_id": item["article_id"],
                "doc_sent_num": item["doc_encodes"].shape[0]
            }
            yield temp_sample

    def input_fn():
        return (tf.data.Dataset.from_generator(
            gen,
            output_types={
                "doc_encodes": tf.float32,
                "article_set": tf.int64,
                "domain1_score": tf.float32,
                "article_id": tf.int64,
                "doc_sent_num": tf.int64
            },
            output_shapes={
                "doc_encodes": [],
                "article_set": [],
                "domain1_score": [None, int(sys_conf["bert_emb_dim"])],
                "article_id": [],
                "doc_sent_num": []
            }
        ))

    return input_fn


def serving_input_receiver_fn():
    """ tensorflow serving的一个输入流函数

    Returns: 略过，自己看

    """
    features = {
        "doc_encodes": tf.placeholder(tf.float32, [1, None, int(sys_conf["bert_emb_dim"])]),
        "prompt_encodes": tf.placeholder(tf.float32, [1, None, int(sys_conf["bert_emb_dim"])]),
        "article_set": tf.placeholder(tf.int64, [None]),
        "domain1_score": tf.placeholder(tf.float32, [None]),
        "article_id": tf.placeholder(tf.int64, [None]),
        "doc_sent_num": tf.placeholder(tf.int64, [None])
    }
    return tf.estimator.export.build_raw_serving_input_receiver_fn(features)


def input_fn_from_tfrecord(tfrecord_path, batch_size, is_training, element_ids):
    """ 以tfrecord为输入，构建该模型需要的input的io

    Args:
        tfrecord_path: tfrecord文件的路径
        batch_size: 模型使用的batch_size,
        is_training: boolean类型标致是否处于训练阶段
        element_ids: 从tfrecord按照element_ids取出对应的所有元素

    Returns: input_fn的handle

    """
    prompts_embedding_path = os.path.join(sys_conf["data_dir"], "prompt.npz")
    if os.path.exists(prompts_embedding_path):
        prompts_embedding = np.load(prompts_embedding_path)["features"][()]
    else:
        raise ValueError("prompts embedding path is not exist, please check")

    features_map = {
        "doc_encodes": tf.VarLenFeature(dtype=tf.float32),
        "doc_encodes_shape": tf.FixedLenFeature(shape=(2,), dtype=tf.int64),
        "article_set": tf.FixedLenFeature(shape=(), dtype=tf.int64),
        "doc_sent_num": tf.FixedLenFeature(shape=(), dtype=tf.int64),
        "domain1_score": tf.FixedLenFeature(shape=(), dtype=tf.float32),
        "article_id": tf.FixedLenFeature(shape=(), dtype=tf.int64)
    }

    def _decode_tfserilized(record, feature_map, shuffle=False):
        example = tf.parse_single_example(record, feature_map)
        temp_example = dict()
        temp_example["doc_encodes"] = tf.sparse_tensor_to_dense(example["doc_encodes"])
        temp_example["doc_encodes"] = tf.reshape(temp_example["doc_encodes"], example["doc_encodes_shape"])
        temp_example["article_set"] = tf.to_int32(example["article_set"])
        temp_example["doc_sent_num"] = tf.to_int32(example["doc_sent_num"])
        temp_example["domain1_score"] = example["domain1_score"]
        temp_example["article_id"] = tf.to_int32(example["article_id"])
        # 只在计算prompt-relevant score的时候有用， 因为在训练prompt-relevant模型的时候是按prompt来训练的，所以把所有sample（包含正负样本）的prompt_encode都赋值一样的
        temp_example["prompt_encodes"] = tf.convert_to_tensor(prompts_embedding[train_conf["prompt_id"]])
        if shuffle:
            if is_training:
                temp_example["doc_encodes"] = tf.random.shuffle(temp_example["doc_encodes"])
                temp_example["article_id"] = temp_example["article_id"] + 100000
            else:
                temp_example["doc_encodes"] = tf.random.shuffle(temp_example["doc_encodes"], seed=1)
                temp_example["article_id"] = temp_example["article_id"] + 100000
        return temp_example

    def input_fn():
        ds = tf.data.TFRecordDataset(tfrecord_path)
        ds1 = ds.map(lambda record: _decode_tfserilized(record, features_map))
        ds2 = ds.map(lambda record: _decode_tfserilized(record, features_map, True))
        ds = ds1.concatenate(ds2)
        ds = ds.filter(predicate=lambda record: tf.math.greater(
            tf.reduce_sum(tf.cast(tf.equal(tf.to_int32(element_ids), record["article_id"]), tf.int32)), 0))
        if is_training:
            ds = ds.repeat()
            ds = ds.shuffle(buffer_size=2000)
        ds = ds.padded_batch(batch_size=batch_size,
                             padded_shapes={
                                 "article_set": [],
                                 "doc_encodes": [None, int(sys_conf["bert_emb_dim"])],
                                 "doc_sent_num": [],
                                 "domain1_score": [],
                                 "article_id": [],
                                 "prompt_encodes": [None, int(sys_conf["bert_emb_dim"])]
                             },
                             drop_remainder=False)
        return ds

    return input_fn


def read_asap_dataset():
    # asap数据集的相关参数，配置，这里做全局变量使用，方便下面三个score predictor调用
    asap_csv_file_path = os.path.join(sys_conf["data_dir"], "training_set_rel3.csv")
    if not os.path.exists(asap_csv_file_path):
        raise ValueError("asap_file_path is invalid.")
    asap_dataset = pd.read_csv(asap_csv_file_path)
    articles_id = list(asap_dataset["essay_id"])
    articles_set = list(asap_dataset["essay_set"])
    domain1_score = asap_dataset["domain1_score"]
    handmark_scores = dict(zip(articles_id, domain1_score))
    set_ids = {
        1: [],
        2: [],
        3: [],
        4: [],
        5: [],
        6: [],
        7: [],
        8: []
    }
    for i in range(len(articles_id)):
        set_ids[articles_set[i]].append(articles_id[i])

    return articles_id, articles_set, set_ids, handmark_scores


def generate_xgboost_train_set(articles_id,
                               articles_set,
                               domain1_scores,
                               train_set_gec_result_path,
                               train_set_saved_path):
    """ 根据训练集gec的结果生成xgboost的训练数据集

    Args:
        articles_id: 训练集文章的id组成的list
        articles_set: 训练集文章的set组成的list
        domain1_scores: 训练集文章的手工标注的分数，因为asap数据集把这个分数称为domain1_scores
        train_set_gec_result_path: 训练集文章过了gec引擎所产生的结果文件的路径，文件形式为一行对应一个文章的gec结果。
        train_set_saved_path: 保存成npz文件类型，npz文件的保存路径

    Returns: 无

    """
    dataset_gec_path = train_set_gec_result_path
    dataset_xgboost_train_file = train_set_saved_path

    # normalized_scores
    handmark_scores = dict(zip(articles_id, domain1_scores))

    # normalized_orgin_scores
    handmark_normalized_scores = {}
    for key, value in handmark_scores.items():
        article_set_id = articles_set[articles_id.index(key)]
        min_value = dataset_score_range[article_set_id][0]
        max_value = dataset_score_range[article_set_id][1]
        normalize_value = (value - min_value) / (max_value - min_value)
        handmark_normalized_scores[key] = normalize_value

    features = {}
    with open(dataset_gec_path, encoding="utf-8") as fr:
        for line in fr:
            line_split = line.split("  ", 1)
            id = int(line_split[0].strip())
            gec_output = json.loads(line_split[1].strip())
            features[id] = Document(gec_output).features

    # TODO(Jiawei): may have bugs if basic_scores的key和features的key不一样
    for key, value in handmark_normalized_scores.items():
        if key in features:
            features[key].append(value)

    np.savez(dataset_xgboost_train_file, features=features)


def sentence_tokenize(documents):
    """分句函数，将一整段文本进行分句

    Args:
        documents: 待分句的document, string类型

    Returns: 句子组成的list

    """
    # 查看
    locations = [-1]
    locations.extend([item.start() for item in re.finditer(r'[\.\?\!](?=[^ \W\d])', documents)])
    locations.append(len(documents))
    sentences = [documents[locations[i] + 1:locations[i + 1] + 1] for i in range(len(locations) - 1)]
    pre_split_documents = " ".join(sentences)

    sentences = nltk.sent_tokenize(pre_split_documents)
    return sentences


if __name__ == "__main__":
    # 使用bert对prompt进行encode
    bc = BertClient()
    result = {}
    prompt_npz = "/Users/liujiawei/Desktop/asap_dataset/prompt.npz"
    with open("/Users/liujiawei/Desktop/asap_dataset/prompt", "r") as reader:
        for i, line in enumerate(reader):
            sentences = sentence_tokenize(line.strip())
            encodes = bc.encode(sentences)
            result[i + 1] = encodes
    np.savez(prompt_npz, features=result)
